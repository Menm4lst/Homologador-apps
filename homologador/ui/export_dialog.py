"""
Sistema de exportación avanzada para homologaciones.
Permite exportar datos en múltiples formatos con opciones personalizadas.
"""

from __future__ import annotations

from datetime import datetime
from typing import Any, Dict, List, Optional, TypedDict
import json
import logging

from PyQt6.QtCore import Qt, QThread, pyqtSignal, pyqtSlot
from PyQt6.QtGui import QFont
import csv
from PyQt6.QtWidgets import (
    QButtonGroup,
    QCheckBox,
    QComboBox,
    QDialog,
    QFileDialog,
    QFormLayout,
    QFrame,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QRadioButton,
    QSpinBox,
    QTabWidget,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)



from .notification_system import send_error, send_success
from .theme import ThemeType, get_current_theme
logger = logging.getLogger(__name__)


Record = Dict[str, Any]


class ExportConfig(TypedDict):
    file_path: str
    format: str
    selected_fields: List[str]
    include_headers: bool
    include_metadata: bool
    date_filter_enabled: bool


class ExportWorker(QThread):
    """Worker thread para exportar datos sin bloquear la UI."""
    
    progress_updated = pyqtSignal(int, str)
    export_completed = pyqtSignal(str)
    export_failed = pyqtSignal(str)

    def __init__(self, data: List[Record], export_config: ExportConfig):
        super().__init__()
        self.data: List[Record] = data
        self.config: ExportConfig = export_config
    
    def run(self):
        """Ejecuta la exportación en segundo plano."""
        try:
            file_path = self.config['file_path']
            export_format = self.config['format']
            
            self.progress_updated.emit(10, "Preparando datos...")
            
            # Filtrar datos según configuración
            filtered_data = self.filter_data()
            
            self.progress_updated.emit(30, "Procesando datos...")
            
            if export_format == 'csv':
                self.export_csv(filtered_data, file_path)
            elif export_format == 'json':
                self.export_json(filtered_data, file_path)
            elif export_format == 'excel':
                self.export_excel(filtered_data, file_path)
            elif export_format == 'pdf':
                self.export_pdf(filtered_data, file_path)
            
            self.progress_updated.emit(100, "Exportación completada")
            self.export_completed.emit(file_path)
            
        except Exception as e:
            logger.error(f"Error en exportación: {e}")
            self.export_failed.emit(str(e))
    
    def filter_data(self) -> List[Record]:
        """Filtra los datos según la configuración."""
        filtered: List[Record] = []
        
        for item in self.data:
            # Aplicar filtros si están configurados
            if self.config.get('date_filter_enabled'):
                # Implementar filtros de fecha si es necesario
                pass
            
            # Seleccionar solo campos requeridos
            filtered_item: Record = {}
            for field in self.config['selected_fields']:
                filtered_item[field] = item.get(field, '')
            
            filtered.append(filtered_item)
        
        return filtered
    
    def export_csv(self, data: List[Record], file_path: str) -> None:
        """Exporta datos a formato CSV."""
        self.progress_updated.emit(50, "Generando archivo CSV...")
        
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            if not data:
                return
            
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            
            if self.config.get('include_headers', True):
                writer.writeheader()
            
            total = len(data)
            for i, row in enumerate(data):
                writer.writerow(row)
                
                # Actualizar progreso
                progress = 50 + int((i / total) * 40)
                self.progress_updated.emit(progress, f"Escribiendo registro {i+1}/{total}")
    
    def export_json(self, data: List[Record], file_path: str) -> None:
        """Exporta datos a formato JSON."""
        self.progress_updated.emit(50, "Generando archivo JSON...")
        
        export_data = {
            'metadata': {
                'export_date': datetime.now().isoformat(),
                'total_records': len(data),
                'exported_by': 'Homologador v1.0.0'
            },
            'data': data
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
        
        self.progress_updated.emit(90, "Archivo JSON generado")
    
    def export_excel(self, data: List[Record], file_path: str) -> None:
        """Exporta datos a formato Excel (requiere openpyxl)."""
        try:
            

            from openpyxl.styles import Alignment, Font, PatternFill
            import openpyxl
            self.progress_updated.emit(50, "Generando archivo Excel...")
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Homologaciones"
            
            if not data:
                wb.save(file_path)
                return
            
            # Escribir encabezados
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Escribir datos
            total = len(data)
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
                
                # Actualizar progreso
                progress = 50 + int(((row_idx - 2) / total) * 40)
                self.progress_updated.emit(progress, f"Escribiendo fila {row_idx-1}/{total}")
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            
        except ImportError:
            raise Exception("Para exportar a Excel, instale: pip install openpyxl")
    
    def export_pdf(self, data: List[Record], file_path: str) -> None:
        """Exporta datos a formato PDF (requiere reportlab)."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
            
            self.progress_updated.emit(50, "Generando archivo PDF...")
            
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            elements: List[Any] = []
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Centrado
            )
            
            # Título
            title = Paragraph("Reporte de Homologaciones", title_style)
            elements.append(title)
            elements.append(Spacer(1, 12))
            
            # Información del reporte
            info_text = (
                f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>"
                f"Total de registros: {len(data)}"
            )
            info = Paragraph(info_text, styles['Normal'])
            elements.append(info)
            elements.append(Spacer(1, 20))
            
            if data:
                # Preparar datos para la tabla
                headers: List[str] = [str(header) for header in data[0].keys()]
                table_data: List[List[str]] = [headers]
                
                for row in data:
                    table_row = [str(row.get(header, ''))[:50] for header in headers]
                    table_data.append(table_row)
                
                # Crear tabla
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(table)
            
            doc.build(elements)
            
        except ImportError:
            raise Exception("Para exportar a PDF, instale: pip install reportlab")


class ExportDialog(QDialog):
    """Dialog para configurar y ejecutar exportaciones."""
    
    def __init__(self, parent: Optional[QWidget] = None, data: Optional[List[Record]] = None):
        super().__init__(parent)
        self.data: List[Record] = data or []
        self.export_worker: Optional[ExportWorker] = None
        self.setup_ui()
        self.apply_theme_styles()
    
    def setup_ui(self):
        """Configura la interfaz del dialog de exportación."""
        self.setWindowTitle("📄 Exportar Datos")
        self.setModal(True)
        self.resize(600, 500)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        # Título
        title = QLabel("Exportación de Homologaciones")
        title_font = QFont()
        title_font.setBold(True)
        title_font.setPointSize(14)
        title.setFont(title_font)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)
        
        # Pestañas para opciones
        tab_widget = QTabWidget()
        
        # Pestaña de formato
        self.setup_format_tab(tab_widget)
        
        # Pestaña de campos
        self.setup_fields_tab(tab_widget)
        
        # Pestaña de opciones
        self.setup_options_tab(tab_widget)
        
        layout.addWidget(tab_widget)
        
        # Barra de progreso
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # Label de estado
        self.status_label = QLabel("")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.status_label)
        
        # Botones
        self.setup_buttons(layout)
    
    def setup_format_tab(self, tab_widget):
        """Configura la pestaña de selección de formato."""
        format_tab = QWidget()
        layout = QVBoxLayout(format_tab)
        
        # Selector de formato
        format_group = QGroupBox("Formato de Exportación")
        format_layout = QVBoxLayout(format_group)
        
        self.format_group = QButtonGroup()
        
        # CSV
        csv_radio = QRadioButton("📊 CSV (Comma Separated Values)")
        csv_radio.setChecked(True)
        csv_radio.setToolTip("Formato compatible con Excel y otras aplicaciones")
        self.format_group.addButton(csv_radio, 0)
        format_layout.addWidget(csv_radio)
        
        # JSON
        json_radio = QRadioButton("🔧 JSON (JavaScript Object Notation)")
        json_radio.setToolTip("Formato estructurado para intercambio de datos")
        self.format_group.addButton(json_radio, 1)
        format_layout.addWidget(json_radio)
        
        # Excel
        excel_radio = QRadioButton("📈 Excel (.xlsx)")
        excel_radio.setToolTip("Formato nativo de Microsoft Excel")
        self.format_group.addButton(excel_radio, 2)
        format_layout.addWidget(excel_radio)
        
        # PDF
        pdf_radio = QRadioButton("📄 PDF (Portable Document Format)")
        pdf_radio.setToolTip("Documento portable para impresión")
        self.format_group.addButton(pdf_radio, 3)
        format_layout.addWidget(pdf_radio)
        
        layout.addWidget(format_group)
        
        # Selector de archivo
        file_group = QGroupBox("Archivo de Destino")
        file_layout = QHBoxLayout(file_group)
        
        self.file_path_edit = QLineEdit()
        self.file_path_edit.setPlaceholderText("Seleccione la ubicación del archivo...")
        file_layout.addWidget(self.file_path_edit)
        
        browse_button = QPushButton("Examinar...")
        browse_button.clicked.connect(self.browse_file)
        file_layout.addWidget(browse_button)
        
        layout.addWidget(file_group)
        
        layout.addStretch()
        
        tab_widget.addTab(format_tab, "📁 Formato")
    
    def setup_fields_tab(self, tab_widget):
        """Configura la pestaña de selección de campos."""
        fields_tab = QWidget()
        layout = QVBoxLayout(fields_tab)
        
        # Información
        info_label = QLabel("Seleccione los campos que desea incluir en la exportación:")
        layout.addWidget(info_label)
        
        # Botones de selección masiva
        buttons_layout = QHBoxLayout()
        
        select_all_btn = QPushButton("Seleccionar Todo")
        select_all_btn.clicked.connect(self.select_all_fields)
        buttons_layout.addWidget(select_all_btn)
        
        select_none_btn = QPushButton("Deseleccionar Todo")
        select_none_btn.clicked.connect(self.select_no_fields)
        buttons_layout.addWidget(select_none_btn)
        
        buttons_layout.addStretch()
        layout.addLayout(buttons_layout)
        
        # Lista de campos
        fields_group = QGroupBox("Campos Disponibles")
        fields_layout = QVBoxLayout(fields_group)
        
        # Definir campos disponibles
        self.field_checkboxes = {}
        available_fields = [
            ('id', 'ID'),
            ('real_name', 'Nombre Real'),
            ('logical_name', 'Nombre Lógico'),
            ('kb_url', 'URL Documentación'),
            ('kb_sync', 'KB Sync'),
            ('homologation_date', 'Fecha Homologación'),
            ('has_previous_versions', 'Versiones Previas'),
            ('repository_location', 'Repositorio'),
            ('details', 'Detalles'),
            ('created_by_username', 'Creado Por'),
            ('created_at', 'Fecha Creación'),
            ('updated_at', 'Fecha Actualización')
        ]
        
        for field_key, field_label in available_fields:
            checkbox = QCheckBox(field_label)
            checkbox.setChecked(True)  # Por defecto todos seleccionados
            self.field_checkboxes[field_key] = checkbox
            fields_layout.addWidget(checkbox)
        
        layout.addWidget(fields_group)
        
        tab_widget.addTab(fields_tab, "📋 Campos")
    
    def setup_options_tab(self, tab_widget):
        """Configura la pestaña de opciones adicionales."""
        options_tab = QWidget()
        layout = QVBoxLayout(options_tab)
        
        # Opciones generales
        general_group = QGroupBox("Opciones Generales")
        general_layout = QVBoxLayout(general_group)
        
        self.include_headers_cb = QCheckBox("Incluir encabezados de columna")
        self.include_headers_cb.setChecked(True)
        general_layout.addWidget(self.include_headers_cb)
        
        self.include_metadata_cb = QCheckBox("Incluir metadatos (fecha, total registros)")
        self.include_metadata_cb.setChecked(True)
        general_layout.addWidget(self.include_metadata_cb)
        
        layout.addWidget(general_group)
        
        # Opciones de filtrado
        filter_group = QGroupBox("Filtros Adicionales")
        filter_layout = QVBoxLayout(filter_group)
        
        self.date_filter_cb = QCheckBox("Filtrar por rango de fechas")
        filter_layout.addWidget(self.date_filter_cb)
        
        # TODO: Agregar más opciones de filtrado según necesidades
        
        layout.addWidget(filter_group)
        
        layout.addStretch()
        
        tab_widget.addTab(options_tab, "⚙️ Opciones")
    
    def setup_buttons(self, layout):
        """Configura los botones del dialog."""
        button_layout = QHBoxLayout()
        
        # Información de registros
        info_label = QLabel(f"Total de registros: {len(self.data)}")
        button_layout.addWidget(info_label)
        
        button_layout.addStretch()
        
        # Botón cancelar
        cancel_button = QPushButton("Cancelar")
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(cancel_button)
        
        # Botón exportar
        self.export_button = QPushButton("🚀 Exportar")
        self.export_button.clicked.connect(self.start_export)
        button_layout.addWidget(self.export_button)
        
        layout.addLayout(button_layout)
    
    def browse_file(self):
        """Abre dialog para seleccionar archivo de destino."""
        format_id = self.format_group.checkedId()
        
        # Configurar filtros según formato
        if format_id == 0:  # CSV
            file_filter = "CSV Files (*.csv)"
            default_ext = ".csv"
        elif format_id == 1:  # JSON
            file_filter = "JSON Files (*.json)"
            default_ext = ".json"
        elif format_id == 2:  # Excel
            file_filter = "Excel Files (*.xlsx)"
            default_ext = ".xlsx"
        elif format_id == 3:  # PDF
            file_filter = "PDF Files (*.pdf)"
            default_ext = ".pdf"
        else:
            file_filter = "All Files (*.*)"
            default_ext = ""
        
        # Generar nombre por defecto
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"homologaciones_{timestamp}{default_ext}"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar archivo de exportación",
            default_name,
            file_filter
        )
        
        if file_path:
            self.file_path_edit.setText(file_path)
    
    def select_all_fields(self):
        """Selecciona todos los campos."""
        for checkbox in self.field_checkboxes.values():
            checkbox.setChecked(True)
    
    def select_no_fields(self):
        """Deselecciona todos los campos."""
        for checkbox in self.field_checkboxes.values():
            checkbox.setChecked(False)
    
    def start_export(self):
        """Inicia el proceso de exportación."""
        # Validar configuración
        if not self.file_path_edit.text().strip():
            QMessageBox.warning(self, "Advertencia", "Seleccione un archivo de destino")
            return
        
        selected_fields: List[str] = [
            field for field, checkbox in self.field_checkboxes.items()
            if checkbox.isChecked()
        ]
        
        if not selected_fields:
            QMessageBox.warning(self, "Advertencia", "Seleccione al menos un campo para exportar")
            return
        
        # Configurar exportación
        format_id = self.format_group.checkedId()
        format_names: List[str] = ['csv', 'json', 'excel', 'pdf']
        format_index = format_id if 0 <= format_id < len(format_names) else 0
        export_format = format_names[format_index]

        export_config: ExportConfig = {
            'file_path': str(self.file_path_edit.text()),
            'format': export_format,
            'selected_fields': selected_fields,
            'include_headers': self.include_headers_cb.isChecked(),
            'include_metadata': self.include_metadata_cb.isChecked(),
            'date_filter_enabled': self.date_filter_cb.isChecked(),
        }
        
        # Deshabilitar interfaz durante exportación
        self.export_button.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        
        # Iniciar worker
        self.export_worker = ExportWorker(self.data, export_config)
        self.export_worker.progress_updated.connect(self.update_progress)
        self.export_worker.export_completed.connect(self.on_export_completed)
        self.export_worker.export_failed.connect(self.on_export_failed)
        self.export_worker.start()
    
    @pyqtSlot(int, str)
    def update_progress(self, progress: int, message: str):
        """Actualiza la barra de progreso."""
        self.progress_bar.setValue(progress)
        self.status_label.setText(message)
    
    @pyqtSlot(str)
    def on_export_completed(self, file_path: str):
        """Maneja la exportación completada."""
        self.progress_bar.setVisible(False)
        self.status_label.setText("")
        self.export_button.setEnabled(True)
        
        send_success("Exportación Exitosa", f"Exportación completada exitosamente:\n{file_path}", "export_system")
        self.accept()
    
    @pyqtSlot(str)
    def on_export_failed(self, error_message: str):
        """Maneja errores en la exportación."""
        self.progress_bar.setVisible(False)
        self.status_label.setText("")
        self.export_button.setEnabled(True)
        
        send_error("Error de Exportación", f"Error en la exportación:\n{error_message}", "export_system")
    
    def apply_theme_styles(self):
        """Aplica estilos según el tema actual."""
        current_theme = get_current_theme()
        
        if current_theme == ThemeType.DARK:
            self.setStyleSheet("""
                QDialog {
                    background-color: #1a1a1a;
                    color: #ffffff;
                }
                QLabel {
                    color: #ffffff;
                    background-color: transparent;
                }
                QGroupBox {
                    font-weight: bold;
                    border: 1px solid #555555;
                    border-radius: 4px;
                    margin-top: 10px;
                    padding-top: 10px;
                    color: #ffffff;
                }
                QPushButton {
                    background-color: #0078d4;
                    color: #ffffff;
                    border: none;
                    border-radius: 4px;
                    padding: 8px 16px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #106ebe;
                }
                QLineEdit {
                    background-color: #2d2d2d;
                    border: 1px solid #555555;
                    border-radius: 4px;
                    padding: 6px;
                    color: #ffffff;
                }
                QTabWidget::pane {
                    border: 1px solid #555555;
                    background-color: #2d2d2d;
                }
                QTabBar::tab {
                    background-color: #3d3d3d;
                    color: #ffffff;
                    padding: 8px 16px;
                    border: 1px solid #555555;
                }
                QTabBar::tab:selected {
                    background-color: #0078d4;
                }
            """)
        else:
            self.setStyleSheet("""
                QDialog {
                    background-color: #ffffff;
                    color: #333333;
                }
                QLabel {
                    color: #333333;
                    background-color: transparent;
                }
                QGroupBox {
                    font-weight: bold;
                    border: 1px solid #d0d0d0;
                    border-radius: 4px;
                    margin-top: 10px;
                    padding-top: 10px;
                    color: #333333;
                }
                QPushButton {
                    background-color: #0078d4;
                    color: #ffffff;
                    border: none;
                    border-radius: 4px;
                    padding: 8px 16px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #106ebe;
                }
                QLineEdit {
                    background-color: #ffffff;
                    border: 1px solid #d0d0d0;
                    border-radius: 4px;
                    padding: 6px;
                    color: #333333;
                }
                QTabWidget::pane {
                    border: 1px solid #d0d0d0;
                    background-color: #ffffff;
                }
                QTabBar::tab {
                    background-color: #f8f8f8;
                    color: #333333;
                    padding: 8px 16px;
                    border: 1px solid #d0d0d0;
                }
                QTabBar::tab:selected {
                    background-color: #0078d4;
                    color: #ffffff;
                }
            """)
    
    def closeEvent(self, event):
        """Limpia recursos al cerrar."""
        if self.export_worker and self.export_worker.isRunning():
            self.export_worker.terminate()
            self.export_worker.wait()
        event.accept()